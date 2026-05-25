from flask import Flask, request, jsonify, send_from_directory
import requests, os, datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet

app      = Flask(__name__)
RAILWAY  = "https://qamra-production.up.railway.app"
BASE     = os.path.dirname(os.path.abspath(__file__))
EXCEL    = os.path.join(BASE, "guests.xlsx")
PDF      = os.path.join(BASE, "guests.pdf")
FACES    = os.path.join(BASE, "faces")

os.makedirs(FACES, exist_ok=True)

HEADERS = ["WhatsApp", "Timestamp", "Photos Found", "Photo URL", "Confidence %", "Face Image"]

# ── Excel init ─────────────────────────────────────────────
def init_excel():
    if os.path.exists(EXCEL):
        return
    wb = Workbook()
    ws = wb.active
    ws.title = "Guests"
    ws.append(HEADERS)
    for col in range(1, len(HEADERS) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font      = Font(bold=True, color="FAF6EC")
        cell.fill      = PatternFill("solid", fgColor="1A1612")
        cell.alignment = Alignment(horizontal="center")
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 60
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 30
    wb.save(EXCEL)

# ── PDF rebuild ────────────────────────────────────────────
def rebuild_pdf():
    try:
        wb   = load_workbook(EXCEL)
        ws   = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            return

        doc    = SimpleDocTemplate(PDF, pagesize=A4,
                                   topMargin=1.5*cm, bottomMargin=1.5*cm,
                                   leftMargin=1*cm, rightMargin=1*cm)
        styles = getSampleStyleSheet()

        elements = [
            Paragraph("QAMRA — Guest Photo Log", styles["Title"]),
            Spacer(1, 0.2*cm),
            Paragraph(f"Generated: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M')}",
                      styles["Normal"]),
            Spacer(1, 0.5*cm),
        ]

        table_data = [[str(c) if c is not None else "" for c in rows[0]]]
        for row in rows[1:]:
            table_data.append([str(c) if c is not None else "" for c in row])

        t = Table(
            table_data,
            repeatRows=1,
            colWidths=[3.2*cm, 3.8*cm, 2.2*cm, 8*cm, 2.2*cm, 3.5*cm],
        )
        t.setStyle(TableStyle([
            ("BACKGROUND",     (0, 0), (-1,  0), colors.HexColor("#1A1612")),
            ("TEXTCOLOR",      (0, 0), (-1,  0), colors.HexColor("#FAF6EC")),
            ("FONTNAME",       (0, 0), (-1,  0), "Helvetica-Bold"),
            ("FONTSIZE",       (0, 0), (-1,  0), 8),
            ("FONTSIZE",       (0, 1), (-1, -1), 7),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1),
             [colors.HexColor("#FAF6EC"), colors.HexColor("#F2EDE3")]),
            ("GRID",           (0, 0), (-1, -1), 0.4, colors.HexColor("#E8E1D2")),
            ("VALIGN",         (0, 0), (-1, -1), "MIDDLE"),
            ("LEFTPADDING",    (0, 0), (-1, -1), 4),
            ("RIGHTPADDING",   (0, 0), (-1, -1), 4),
            ("TOPPADDING",     (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING",  (0, 0), (-1, -1), 3),
            ("WORDWRAP",       (3, 1), (3,  -1), True),
        ]))
        elements.append(t)
        doc.build(elements)
    except Exception as e:
        print(f"[PDF] Error: {e}")

init_excel()

# ── Routes ─────────────────────────────────────────────────
@app.route("/")
def index():
    return send_from_directory(BASE, "index.html")

@app.route("/<path:path>")
def static_file(path):
    return send_from_directory(BASE, path)

@app.route("/api/match", methods=["POST"])
def proxy_match():
    f = request.files.get("photo")
    if not f:
        return jsonify({"error": "no photo"}), 400
    photo_bytes = f.read()

    # Save selfie for post-event re-search
    ts        = datetime.datetime.now().strftime("%Y%m%d_%H%M%S_%f")
    face_name = f"face_{ts}.jpg"
    with open(os.path.join(FACES, face_name), "wb") as fp:
        fp.write(photo_bytes)

    try:
        r = requests.post(
            f"{RAILWAY}/match",
            files={"photo": (f.filename or "selfie.jpg", photo_bytes,
                             f.content_type or "image/jpeg")},
            timeout=30,
        )
        resp = r.json()
        resp["face_path"] = face_name
        return jsonify(resp), r.status_code
    except requests.exceptions.Timeout:
        return jsonify({"error": "timeout", "message": "انتهت مهلة الاتصال"}), 504
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/folder-status/<session_id>", methods=["GET"])
def proxy_folder_status(session_id):
    try:
        r = requests.get(f"{RAILWAY}/folder-status/{session_id}", timeout=10)
        return jsonify(r.json()), r.status_code
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/log-guest", methods=["POST"])
def log_guest():
    data      = request.json or {}
    whatsapp  = data.get("whatsapp", "")
    photos    = data.get("photos", [])      # [{url, confidence}, ...]
    face_path = data.get("face_path", "")
    now       = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    try:
        wb = load_workbook(EXCEL)
        ws = wb.active

        if not photos:
            ws.append([whatsapp, now, 0, "", "", face_path])
        else:
            for i, photo in enumerate(photos):
                ws.append([
                    whatsapp    if i == 0 else "",
                    now         if i == 0 else "",
                    len(photos) if i == 0 else "",
                    photo.get("url", ""),
                    round(photo.get("confidence", 0)),
                    face_path   if i == 0 else "",
                ])

        wb.save(EXCEL)
        rebuild_pdf()
        return jsonify({"status": "logged", "rows": max(len(photos), 1)}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

if __name__ == "__main__":
    print("\nQamra Kiosk -- http://localhost:5000\n")
    app.run(host="0.0.0.0", port=5000, debug=False)
